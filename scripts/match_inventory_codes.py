import csv
import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


CATALOG = Path("data/extraccion-catalogo/precios_auditados.csv")
OUT_DIR = Path("data/extraccion-catalogo")
INVENTORY = next(Path(".").glob("Inventario sede rionegro*.XLSX"))


STOPWORDS = {
    "x", "de", "la", "el", "y", "en", "con", "sin", "und", "un", "u", "unidad",
    "unidades", "bs", "bol", "bolsa", "bolsas", "pqt", "pqt.", "sob", "sobre",
    "sobres", "caja", "cajax", "disp", "plex", "pleg", "tar", "tc", "dx"
}

REPLACEMENTS = {
    "gta": "galleta",
    "gta.": "galleta",
    "gallet": "galleta",
    "chocol": "chocolate",
    "clas": "clasico",
    "clasica": "clasico",
    "cyc": "canela clavos",
    "fco": "frasco",
    "fco.": "frasco",
    "vlla": "vainilla",
    "vlla.": "vainilla",
    "cappno": "capuccino",
    "cappuccino": "capuccino",
    "spaguetti": "spaghetti",
    "comarri": "comarrico",
    "zenu": "zenu",
    "zenú": "zenu",
    "colcafé": "colcafe",
    "pińa": "pina",
    "champiñon": "champinon",
    "champińon": "champinon",
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def normalize(value: str) -> str:
    text = strip_accents(str(value).lower())
    text = text.replace("&", " ")
    text = re.sub(r"(\d+)\s*(g|gr|kg|ml|lt|l|und|un|sob|pqt|fco|tar|bol|plex|pleg)", r"\1 \2", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    tokens = []
    for token in text.split():
        token = REPLACEMENTS.get(token, token)
        if token not in STOPWORDS:
            tokens.extend(token.split())
    return " ".join(tokens)


def tokens(value: str) -> set[str]:
    return set(normalize(value).split())


def quantities(value: str) -> set[str]:
    text = strip_accents(str(value).lower())
    matches = re.findall(r"\d+(?:[.,]\d+)?\s*(?:g|gr|kg|ml|lt|l|und|un|sob|pqt|fco|tar|bol|plex|pleg)", text)
    return {re.sub(r"\s+", "", match.replace(",", ".")) for match in matches}


def load_catalog() -> list[dict[str, str]]:
    with CATALOG.open(encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def load_inventory() -> list[dict[str, str]]:
    workbook = load_workbook(INVENTORY, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(value).strip() for value in rows[0]]
    result = []
    for row in rows[1:]:
        item = dict(zip(header, row))
        code = item.get("Material")
        name = item.get("Texto breve de material")
        if code is None or not name:
            continue
        result.append({
            "codigo_inventario": str(int(code)) if isinstance(code, float) else str(code).strip(),
            "nombre_inventario": str(name).strip(),
            "centro": str(item.get("Centro") or "").strip(),
            "unidad": str(item.get("Unidad medida base") or "").strip(),
            "disponible": str(item.get("Disponible") or "").strip(),
        })
    return result


def score(catalog_name: str, inventory_name: str) -> tuple[float, str]:
    catalog_norm = normalize(catalog_name)
    inventory_norm = normalize(inventory_name)
    catalog_tokens = set(catalog_norm.split())
    inventory_tokens = set(inventory_norm.split())
    if not catalog_tokens or not inventory_tokens:
        return 0, ""

    overlap = len(catalog_tokens & inventory_tokens) / len(catalog_tokens | inventory_tokens)
    seq = SequenceMatcher(None, catalog_norm, inventory_norm).ratio()
    catalog_qty = quantities(catalog_name)
    inventory_qty = quantities(inventory_name)

    qty_bonus = 0
    if catalog_qty and inventory_qty:
      qty_bonus = len(catalog_qty & inventory_qty) / max(len(catalog_qty), len(inventory_qty))

    value = (overlap * 0.48) + (seq * 0.32) + (qty_bonus * 0.20)
    detail = f"tokens={overlap:.2f};texto={seq:.2f};cantidades={qty_bonus:.2f}"
    return round(value * 100, 2), detail


def write_csv(path: Path, rows: list[dict], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    catalog = load_catalog()
    inventory = load_inventory()
    inventory_by_code = {row["codigo_inventario"]: row for row in inventory}

    suggestions = []
    high_confidence = []
    catalog_code_missing = []

    for product in catalog:
        ranked = []
        for item in inventory:
            value, detail = score(product["nombre"], item["nombre_inventario"])
            if value >= 45:
                ranked.append((value, detail, item))
        ranked.sort(key=lambda item: item[0], reverse=True)
        best = ranked[:5]
        exact_code = inventory_by_code.get(product["codigo"])

        if product["codigo"] not in inventory_by_code:
            catalog_code_missing.append({
                "codigo_catalogo": product["codigo"],
                "nombre_catalogo": product["nombre"],
                "marca": product["marca"],
            })

        second_score = best[1][0] if len(best) > 1 else 0
        for rank, (value, detail, item) in enumerate(best, start=1):
            row = {
                "codigo_catalogo": product["codigo"],
                "nombre_catalogo": product["nombre"],
                "marca": product["marca"],
                "precio_catalogo": product.get("precio", ""),
                "codigo_inventario": item["codigo_inventario"],
                "nombre_inventario": item["nombre_inventario"],
                "disponible": item["disponible"],
                "confianza": value,
                "detalle": detail,
                "ranking": rank,
                "codigo_actual_existe_en_inventario": "si" if exact_code else "no",
            }
            suggestions.append(row)
            margin = value - second_score
            if (
                rank == 1
                and not exact_code
                and value >= 88
                and margin >= 5
                and product["codigo"] != item["codigo_inventario"]
            ):
                row["margen_vs_segunda_opcion"] = round(margin, 2)
                high_confidence.append(row)

    fields = [
        "codigo_catalogo", "nombre_catalogo", "marca", "precio_catalogo",
        "codigo_inventario", "nombre_inventario", "disponible", "confianza",
        "detalle", "ranking", "codigo_actual_existe_en_inventario",
    ]
    write_csv(OUT_DIR / "sugerencias_codigos_inventario.csv", suggestions, fields)
    write_csv(OUT_DIR / "actualizaciones_alta_confianza.csv", high_confidence, fields)
    write_csv(
        OUT_DIR / "codigos_catalogo_no_en_inventario.csv",
        catalog_code_missing,
        ["codigo_catalogo", "nombre_catalogo", "marca"],
    )
    write_csv(
        OUT_DIR / "inventario_normalizado.csv",
        inventory,
        ["codigo_inventario", "nombre_inventario", "centro", "unidad", "disponible"],
    )

    print(f"Productos catalogo: {len(catalog)}")
    print(f"Productos inventario: {len(inventory)}")
    print(f"Sugerencias generadas: {len(suggestions)}")
    print(f"Actualizaciones alta confianza: {len(high_confidence)}")
    print(f"Codigos del catalogo no presentes en inventario: {len(catalog_code_missing)}")


if __name__ == "__main__":
    main()
