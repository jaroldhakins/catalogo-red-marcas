from pathlib import Path

from openpyxl import load_workbook


workbook_path = next(Path(".").glob("Inventario sede rionegro*.XLSX"))
workbook = load_workbook(workbook_path, data_only=True, read_only=True)

print(f"Archivo: {workbook_path}")
print(f"Hojas: {workbook.sheetnames}")

for sheet in workbook.worksheets:
    print(f"\n--- {sheet.title} ---")
    print(f"Filas: {sheet.max_row} Columnas: {sheet.max_column}")
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [value for value in row if value not in (None, "")]
        if values:
            print(f"Fila {row_index}: {values[:12]}")
        if row_index >= 12:
            break
